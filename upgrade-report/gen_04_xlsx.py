# -*- coding: utf-8 -*-
"""
生成《04-功能对标分析.xlsx》
- 使用 Python 3 + openpyxl（系统已安装 3.1.5）
- 数据来源：2026-08 通过 WebSearch 搜索 GitHub 真实开源项目（非虚构）
- Star 数标注“约 xx k / 未查到”的，均来自搜索结果原文，未逐一调用 GitHub API 复核，
  可能与实时数据有偏差，仅作量级参考。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# (项目名, Stars, GitHub 地址, 核心功能, 我们是否已有, 功能成熟度, 是否值得引入)
ROWS = [
    # ---------- A 类：本地 ChatGPT / LLM 桌面客户端 ----------
    ("NextChat（ChatGPT-Next-Web）", "约 81k~100k+（2025-02 报道 80.7k）",
     "https://github.com/ChatGPTNextWeb/NextChat",
     "多模型聚合 Web/PWA/桌面客户端：OpenAI/Claude/Gemini/DeepSeek/本地模型，自定义模型列表、上下文自动压缩、离线缓存、Tauri 桌面壳、多语言",
     "部分", "成熟", "参考借鉴"),
    ("LobeChat", "约 45k~68k（2025-12 报道 67.8k）",
     "https://github.com/lobehub/lobe-chat",
     "高颜值 LLM 聊天框架：TTS/STT 语音会话、Function Calling 插件系统、助手市场、文生图、PWA、多主题、i18n 自动化",
     "部分", "成熟", "值得引入"),
    ("Chatbox", "约 30.7k（2025-02 报道）",
     "https://github.com/Bin-Huang/chatbox",
     "跨平台 OpenAI API 桌面客户端：本地存储聊天记录、多会话、提示词库、消息引用、流式回复、图像生成、团队共享 API",
     "部分", "成熟", "值得引入"),
    ("lencx/ChatGPT（ChatGPT 桌面客户端）", "约 52.7k（2026-02 报道）",
     "https://github.com/lencx/ChatGPT",
     "Tauri 跨平台 ChatGPT 桌面壳：对话导出 Markdown/PNG/PDF、提示词库、全局快捷键、本地文件交互、旧系统兼容",
     "无", "成熟", "值得引入"),
    ("Open WebUI", "约 74.6k（2025-02 报道）",
     "https://github.com/open-webui/open-webui",
     "自托管 AI 平台：Ollama/OpenAI 接入、RAG 文档上传、多用户 RBAC、对话分支、Docker 一键部署",
     "无", "成熟", "不引入"),
    ("AnythingLLM", "约 63.9k（star-history 2026-07）",
     "https://github.com/Mintplex-Labs/anything-llm",
     "本地优先 RAG 桌面应用：多格式文档转上下文、向量库（Qdrant/Weaviate）、多 LLM 接入、知识空间与权限",
     "无", "成熟", "参考借鉴"),
    ("LibreChat", "约 20k+（2025-12 报道）",
     "https://github.com/danny-avila/LibreChat",
     "多 provider 聊天：OpenAI/Azure/Anthropic/Google 等 15+、ChatGPT 兼容插件、对话搜索、多用户权限、审计日志",
     "无", "成熟", "不引入"),
    ("Ollama", "约 162k~176k（2026-02 报道）",
     "https://github.com/ollama/ollama",
     "本地 LLM 运行器：一键运行 Llama/Qwen/DeepSeek 等开源模型，OpenAI 兼容 API",
     "无", "成熟", "参考借鉴"),
    ("Page Assist", "约 4.7k（2025-02 报道）",
     "https://github.com/n4ze3m/page-assist",
     "浏览器本地 AI 侧边栏插件：对接 Ollama 等本地模型，页面划词问答",
     "无", "成长中", "不引入"),
    ("OpenAI Codex CLI", "stars 未查到（官方开源，2025-04 发布）",
     "https://github.com/openai/codex",
     "官方开源终端编码 Agent：自然语言开发、多模态截图输入、沙盒、Codex App Server 协议（本项目 app-server 模式直接对接）",
     "有", "成熟", "持续跟踪"),
    # ---------- B 类：手机远程控制桌面/浏览器桥接 ----------
    ("RustDesk", "约 60k~117k（官方 2026-04 称 60k+，2026-07 报道 117k）",
     "https://github.com/rustdesk/rustdesk",
     "开源远程桌面：跨平台、屏幕共享+鼠标键盘控制、文件传输、自建中继、端到端加密（NaCl）",
     "无", "成熟", "参考借鉴"),
    ("Codex-Mini（上游项目）", "stars 未查到",
     "https://github.com/CoimgRain/Codex-Mini",
     "本项目上游：本地桥接思路，手机控制电脑上的 Codex 对话（ChatGPT Win 已基于其思路重写 Windows 自动化层并新增 Android 客户端）",
     "有", "成长中", "持续跟踪"),
    ("BilldDesk", "stars 未查到（开发中项目）",
     "https://github.com/galaxy-s10/billd-desk",
     "WebRTC 开源远程桌面：桌面/网页/Android 互控、屏幕墙、多显示器、文件传输、隐私屏幕，Vue3+WebRTC+Electron+Flutter",
     "无", "早期", "参考借鉴"),
    ("Mobile Remote PC Control", "stars 未查到",
     "https://github.com/smilexizheng/mobile-pc-control-server",
     "手机 Web 远程控制 Windows：文本输入、鼠标移动/点击/拖拽、快捷键面板、自动化组合键，Node.js 服务端",
     "部分", "成长中", "值得引入"),
    ("Kontroller", "stars 未查到",
     "GitHub 搜索「Kontroller」（蓝牙 HID，Android 9+）",
     "Android 手机通过蓝牙 HID 模拟鼠标/键盘，控制电脑/Mac/电视/iPad，无需局域网",
     "无", "早期", "不引入"),
    ("OpenClaw", "报道约 35 万（2026-05 报道口径，未经独立核实）",
     "https://github.com/openclaw/openclaw",
     "自托管 AI Agent 消息桥：接入 WhatsApp/Telegram/Slack 等 20+ 平台，手机发消息指挥本地 AI 执行任务",
     "部分", "成长中", "参考借鉴"),
    # ---------- C 类：定时任务调度 + 队列系统 ----------
    ("Agenda", "约 9.3k（gitmemories）",
     "https://github.com/agenda/agenda",
     "基于 MongoDB 的 Node 任务调度：cron/人类可读间隔、持久化、REST API、重试",
     "部分", "成熟", "不引入"),
    ("Agendash", "约 774",
     "https://github.com/agenda/agendash",
     "Agenda 的 Web 管理界面：任务新增/移除/状态查看",
     "无", "成长中", "不引入"),
    ("node-cron", "stars 未查到",
     "https://github.com/node-cron/node-cron",
     "轻量 cron 调度库：5/6 字段表达式、时区、start/stop",
     "部分", "成熟", "参考借鉴"),
    ("node-schedule", "stars 未查到",
     "https://github.com/node-schedule/node-schedule",
     "复杂调度：cron/日期/RecurrenceRule 规则、时间轮算法",
     "部分", "成熟", "不引入"),
    ("cron-parser", "stars 未查到",
     "https://github.com/harrisiirak/cron-parser",
     "零依赖 cron 解析器：支持标准 + 扩展修饰符（L/W/#/H）、时区、浏览器/Node",
     "部分", "成熟", "参考借鉴"),
    ("Bull", "stars 未查到（约 1.5 万量级）",
     "https://github.com/OptimalBits/bull",
     "Redis 分布式任务队列：延迟任务、cron 重复任务、限速、重试、优先级、并发、崩溃恢复",
     "无", "成熟", "不引入"),
    ("BullMQ", "stars 未查到",
     "https://github.com/taskforcesh/bullmq",
     "Bull 的 TypeScript 重写版：Redis 队列、父子任务、流式进度、UI",
     "无", "成熟", "不引入"),
    # ---------- D 类：Chrome DevTools Protocol 控制工具 ----------
    ("Puppeteer", "约 89k+（2026 调研报道）",
     "https://github.com/puppeteer/puppeteer",
     "Google 官方 CDP 自动化库：无头/有头 Chrome 控制、截图、PDF、性能分析、扩展自动化",
     "无", "成熟", "参考借鉴"),
    ("Playwright", "约 70k+（2026 调研报道）",
     "https://github.com/microsoft/playwright",
     "跨浏览器自动化：Chromium/Firefox/WebKit、自动等待、Trace 录制、网络拦截、MCP 支持",
     "无", "成熟", "参考借鉴"),
    ("Selenium", "约 31k+（2026 调研报道）",
     "https://github.com/SeleniumHQ/selenium",
     "WebDriver 浏览器自动化标准：20 年老牌、多语言绑定",
     "无", "成熟", "不引入"),
    ("chrome-remote-interface", "stars 未查到",
     "https://github.com/cyrus-and/chrome-remote-interface",
     "Node CDP 协议客户端封装：直接连 9222 端口 WebSocket，方法/事件/回调",
     "无", "成熟", "参考借鉴"),
    # ---------- E 类：AI 平台 / 工作流 / Agent ----------
    ("Dify", "约 139k~149k（2026-07 报道）",
     "https://github.com/langgenius/dify",
     "开源 LLM 应用开发平台：可视化工作流、RAG、Agent、模型路由、多用户、LLMOps",
     "无", "成熟", "不引入"),
    ("FastGPT", "约 29k（2026-07 报道）",
     "https://github.com/labring/FastGPT",
     "知识库问答系统：数据处理、Flow 可视化工作流、模型调用",
     "无", "成熟", "不引入"),
    ("n8n", "约 110k~196k（2026 报道 11 万+ / 19.6 万口径不一）",
     "https://github.com/n8n-io/n8n",
     "工作流自动化：400+ 集成节点、Webhook/定时触发、AI Agent、自托管",
     "无", "成熟", "参考借鉴"),
    ("AutoGPT", "约 177k（2025-08 报道）",
     "https://github.com/Significant-Gravitas/AutoGPT",
     "自主 Agent 平台：低代码工作流、持久 Agent、Webhook 触发、监控分析",
     "无", "成熟", "不引入"),
    ("ChatGPT Box", "stars 未查到",
     "https://github.com/josStorer/chatGPTBox",
     "浏览器 ChatGPT 插件：划词翻译、侧边栏对话、截图、多 API（OpenAI/Bing/Poe/Azure）",
     "无", "成长中", "不引入"),
    ("OpenAI Translator", "stars 未查到",
     "https://github.com/yetone/openai-translator",
     "浏览器翻译插件：55 种语言翻译/润色/总结、TTS、截图翻译、生词本",
     "无", "成熟", "参考借鉴"),
]

def main():
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "04-功能对标分析.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "功能对标"

    header = ["序号", "项目名", "Stars", "项目地址", "核心功能", "我们是否已有", "功能成熟度", "是否值得引入"]
    widths = [6, 30, 26, 44, 70, 12, 12, 14]

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, (title, width) in enumerate(zip(header, widths), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = width

    for i, row in enumerate(ROWS, start=1):
        values = [i, row[0], row[1], row[2], row[3], row[4], row[5], row[6]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = border
            cell.alignment = wrap if col in (2, 3, 4, 5) else center
        # 标记“值得引入”的行
        if row[6] == "值得引入":
            for col in range(1, 9):
                ws.cell(row=i + 1, column=col).fill = PatternFill("solid", fgColor="E2EFDA")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(ROWS) + 1}"

    # 附注页
    ws2 = wb.create_sheet("数据说明")
    notes = [
        "数据来源与口径：",
        "1. 全部项目名与 GitHub 地址来自 2026-08 通过 WebSearch 检索到的真实搜索结果，未虚构项目。",
        "2. Star 数为搜索结果中报道/收录的口径（标注了报道时间），未逐一调用 GitHub API 复核，仅作量级参考；",
        "   标注「stars 未查到」表示搜索结果未返回具体数字。",
        "3. 「我们是否已有」基于对 ChatGPT Win（v4.0.0）源码的核对：有=已具备同等能力；",
        "   部分=已有近似/雏形能力；无=未实现。",
        "4. 搜索数量说明：本表共收录 33 个真实存在的开源项目（目标 20~40），覆盖四个方向：",
        "   A 类本地 LLM 客户端 10 个、B 类手机远程控制桥接 6 个、C 类调度/队列 7 个、D 类 CDP 工具 4 个、E 类 AI 平台 6 个。",
        "5. 部分项目（BilldDesk、Kontroller、Mobile Remote PC Control 等）规模较小或处于开发期，star 数未在搜索结果中披露。",
    ]
    for i, text in enumerate(notes, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.alignment = wrap
    ws2.column_dimensions["A"].width = 120

    wb.save(out_path)
    print(f"OK: {out_path}")
    print(f"项目数量: {len(ROWS)}")

if __name__ == "__main__":
    main()
